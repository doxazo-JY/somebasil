"""레시피 템플릿 v5 — v4 + 점장 입력 가이드 보강.

변경점 (v4 → v5):
- README 맨 위에 '받으면 이렇게 시작하세요' 5단계 추가
- ❓ 두 종류로 시각 분리:
  • 진노랑 (필수) — 잔당 사용량/산출량/가격 (없으면 원가 계산 불가)
  • 연회색 (선택) — 정책 확인·이름 매칭 메타 정보
- 샘플 가짜 숫자: 자몽청·말차크림 한 세트씩 채워서 패턴 예시
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\PC\Desktop\somebasil\notes\레시피_템플릿_v5.xlsx"
MENU_DATA = r"C:\Users\PC\Desktop\somebasil\notes\_menu_data.json"
UNK = "❓"
EFF_DATE = "2025-12-15"
ADD_ROWS = 8

CASH_INGREDIENTS = {"원두", "우유", "말차파우더", "잠봉뵈르햄"}

with open(MENU_DATA, encoding="utf-8") as f:
    md = json.load(f)
products = md["products"]
cat_by_name = md["category_by_name"]

# ─── 스타일 ───
HEAD_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
HEAD_FILL = PatternFill("solid", fgColor="1A5C3A")
BODY_FONT = Font(name="맑은 고딕", size=10)
UNK_FILL_REQUIRED = PatternFill("solid", fgColor="FFE066")  # 진노랑 — 필수
UNK_FILL_OPTIONAL = PatternFill("solid", fgColor="EFEFEF")  # 연회색 — 선택
NEW_MENU_FILL = PatternFill("solid", fgColor="E8F1ED")
CASH_FILL = PatternFill("solid", fgColor="E8F1ED")
ADD_FILL = PatternFill("solid", fgColor="FCEAEA")
SAMPLE_FILL = PatternFill("solid", fgColor="EBF4FF")  # 옅은 파랑 — 샘플
SAMPLE_FONT = Font(name="맑은 고딕", size=10, color="888888", italic=True)
THIN = Side(border_style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def style_body(ws, start_row=2, name_col=1, memo_col=None):
    """❓ 셀을 위치 기반으로 색칠.
    - 비고(memo_col)의 ❓는 선택 (회색)
    - 그 외 컬럼의 ❓는 필수 (진노랑)
    """
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = LEFT if cell.column == name_col else CENTER
            v = cell.value
            # ❓ 직접 매칭 (필수)
            if v == UNK:
                cell.fill = UNK_FILL_REQUIRED
            # 비고 셀이고 내용에 ❓ 포함 (선택)
            elif memo_col and cell.column == memo_col and isinstance(v, str) and UNK in v:
                cell.fill = UNK_FILL_OPTIONAL

def autosize(ws, padding=4, max_w=50):
    for col in ws.columns:
        m = 0
        letter = get_column_letter(col[0].column)
        for c in col:
            v = "" if c.value is None else str(c.value)
            l = sum(2 if ord(ch) > 127 else 1 for ch in v)
            if l > m: m = l
        ws.column_dimensions[letter].width = min(m + padding, max_w)


wb = Workbook()

# ═══════════════════════════════════════════════
# 시트 0: README
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = "README"
ws["A1"] = "썸바실 레시피·원가 입력 양식 (v5)"
ws["A1"].font = Font(name="맑은 고딕", bold=True, size=15, color="1A5C3A")
ws.row_dimensions[1].height = 28

guide = [
    "",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "■ 받으면 이렇게 시작하세요 (예상 1.5~2시간)",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "",
    "  ① 시트1 매입재료 가격 채우기 (30분)",
    "     - 진노랑 ❓ 칸에 (1세트 용량 / 세트 수 / 총가격) 입력",
    "     - 영수증·통장 거래내역 보면서. 단가는 자동 계산됨",
    "     - 예: 원두 1kg봉지 2개=70,000원 → 1000 / 2 / 70000",
    "",
    "  ② 시트3 메뉴별 잔당 사용량 채우기 (30~40분)",
    "     - 진노랑 ❓ 칸에 메뉴 1잔에 들어가는 재료 g/펌프/개 입력",
    "     - 매장 매뉴얼 그대로. PDF에서 채운 부분은 확인만",
    "     - 연두색 행 = PDF 미수록 신메뉴, 새로 적어야 함",
    "",
    "  ③ 시트2 청·크림류 1배치 채우기 (20분)",
    "     - 자몽청·딸기청·말차크림 등 한 번 만들 때 N g 만들고",
    "       그 안에 원물 N g + 설탕 N g 들어가는지",
    "     - 자몽청·말차크림 행에 '(예시)' 샘플이 옅은 파랑으로 채워져 있음. 패턴 참고",
    "",
    "  ④ 시트2 베이커리·디저트 채우기 (20분)",
    "     - 베이글·스콘·마들렌·호두파이·초코쿠키·바스크치즈케익·드립백·더치원액",
    "     - '한 번에 N개 만들고 그 때 밀가루 N g + 버터 N g + ...'",
    "     - 정확히 모르면 대략으로라도. 비워두면 그 메뉴 원가 계산 불가",
    "",
    "  ⑤ (시간 되면) 회색 ❓ 메모 + 자유 추가",
    "     - 회색은 정책 확인 / 이름 매칭 등 보조 정보 — 없어도 원가는 나옴",
    "     - 분홍색 빈 행은 빠진 재료 자유 추가 영역",
    "",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "■ 색깔 의미",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "  • 진노랑 ❓     : ★ 필수 — 채우지 않으면 그 메뉴 원가 계산 불가",
    "  • 연회색 ❓ 텍스트: 선택 — 정책 확인·이름 매칭 등 보조 정보",
    "  • 옅은 파랑     : 예시 샘플 (자몽청·말차크림) — 실제 값으로 교체",
    "  • 연두색 행 (시트3): DB에 있는데 PDF 미수록 — 새로 채워야 할 메뉴",
    "  • 연두색 셀 (시트1): 현금 결제 4종 — 거래 단위 추적 가능 그룹",
    "  • 분홍색 행 (시트1): 점장 자유 추가 영역 (빈 행)",
    "",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "■ 시트 구성",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "  1. 재료 — 매입재료/수제재료 마스터",
    "  2. 서브레시피 — 수제재료(청·크림·디저트·드립백 등) 1배치 배합",
    "  3. 레시피 — DB 활성 메뉴 66개 + 잔당 사용량",
    "  4. 포장세트 — 카테고리×온도별 포장재 묶음",
    "",
    "■ 결제방식 컬럼 (시트1)",
    "  • 현금: 통장 거래 단위로 찍히는 재료 (원두·우유·말차파우더·잠봉뵈르햄 4종)",
    "  • 카드: 쿠팡·이마트·씨유 등 묶음 결제",
    "",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "■ 누락 재료/메뉴 직접 추가하는 방법",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "  ① 시트1 매입재료 추가",
    "     - 분홍색 빈 행에 (재료명 / 사용단위 / 종류='매입' / 결제방식 / 1세트용량 / 세트수 / 총가격)",
    "     - 단가 수식은 분홍 행에 미리 들어가 있음",
    "",
    "  ② 시트1 수제재료 추가",
    "     - 분홍색 빈 행에 입력. 종류='수제'로",
    "     - 1세트용량/세트수/총가격은 '-' 그대로, 단가에 '(서브레시피 자동계산)' 입력",
    "     - 시트2 서브레시피에 1배치 산출량 + 구성재료 추가",
    "",
    "  ③ 시트3 메뉴/재료 추가",
    "     - 빈 행에 메뉴명 / 카테고리 / 재료명 / 사용량",
    "     - 한 메뉴에 재료 여러 개면 메뉴명 반복해서 행 여러 개로",
    "",
    "  ④ ★ 재료명은 시트1과 글자 그대로 일치해야 함",
    "     - 띄어쓰기·괄호 차이도 매칭 안됨",
    "     - 새 재료는 시트1에 먼저 등록 후 시트3에서 사용",
    "",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "■ 입력 단위 안내",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "  • g / ml: 무게/부피 그대로",
    "  • 펌프: 시럽 펌프 1회 (디스펜서 한 번 누름)",
    "  • 회: 토핑 1회 분사",
    "  • 개 / 조각: 그대로",
    "  • 사용량 단위 = 시트1의 '사용단위' 컬럼과 일치 (시트3·4에서)",
    "",
    "■ HOT/ICE·매장/포장 차이",
    "  • HOT/ICE는 한 레시피로 통합 (그람수 동일 가정 — PDF 기준)",
    "  • 매장/포장 사용량 다르면 '매장 HOT 기준'으로 입력 + 비고에 차이 메모",
    "",
    "■ 적용시작일",
    "  • 모든 행 = 2025-12-15 통일 (시스템 도입 기준일)",
    "  • 추후 가격 변동 시 새 행 추가 + 그 변경일 입력",
]
for i, line in enumerate(guide, start=2):
    c = ws.cell(row=i, column=1, value=line)
    c.font = Font(name="맑은 고딕", size=11)
ws.column_dimensions["A"].width = 105


# ═══════════════════════════════════════════════
# 시트 1: 재료
# ═══════════════════════════════════════════════
ws = wb.create_sheet("재료")
ws.append([
    "재료명", "사용단위", "종류", "결제방식",
    "1세트 용량", "세트 수", "총 가격(원)", "단가(자동)",
    "적용시작일", "비고",
])

purchased = [
    ("원두", "g", "예: 1kg봉지×2개=70000원 → 1세트용량 1000, 세트수 2"),
    ("핸드드립 원두", "g", "❓일반 원두와 별도 구매? 점장 확인"),
    ("우유", "g", "PDF 기준 g (스팀 우유 무게)"),
    ("바닐라시럽", "펌프", "예: 1병=200펌프 환산"),
    ("초코시럽", "g", ""),
    ("카라멜드리즐", "회", "토핑 1회분"),
    ("시나몬파우더", "회", "토핑"),
    ("코코아파우더", "회", "토핑"),
    ("초코드리즐", "회", "토핑"),
    ("휘핑크림", "g", ""),
    ("흑임자", "g", ""),
    ("생크림", "g", ""),
    ("말차파우더", "g", ""),
    ("순수", "g", "❓PDF '순수'의 정체 확인 필요"),
    ("복숭아시럽", "펌프", ""),
    ("레몬시럽", "펌프", ""),
    ("슈가시럽", "펌프", ""),
    ("라임시럽", "펌프", ""),
    ("TWG 레이디그레이 티백", "개", "종류별 단가 다름"),
    ("TWG 캐모마일 티백", "개", ""),
    ("TWG 페퍼민트 티백", "개", ""),
    ("팔레데떼 티백", "개", "블루오브런던용"),
    ("로얄밀크티 티백", "개", ""),
    ("탄산수", "ml", ""),
    ("아이스티 베이스", "g", "❓시럽? 분말? 매입 형태 확인. 망고/복숭아아이스티 공통"),
    ("자몽", "g", ""),
    ("레몬", "g", "원물 + 가니쉬용"),
    ("사과", "g", "청 + 슬라이스"),
    ("복숭아", "g", "❓도원결의청용 추정"),
    ("패션푸르츠", "g", ""),
    ("딸기", "g", "봄딸기주스에 직접 사용"),
    ("오미자", "g", "오미자청용 원물"),
    ("설탕", "g", "청 + 베이커리"),
    ("잠봉뵈르햄", "g", ""),
    ("버터", "조각", "PDF: 3조각"),
    ("할라피뇨", "개", ""),
    ("어니언드레싱", "펌프", "PDF '한바퀴' = 1펌프 추정"),
    ("선드라이토마토", "개", ""),
    ("발사믹", "펌프", ""),
    ("루꼴라", "g", ""),
    ("크림치즈", "g", "플레인 베이글 + 바스크 치즈케익 베이스"),
    ("마스카포네", "g", ""),
    ("밀가루", "g", ""),
    ("계란", "개", ""),
    ("호두", "g", ""),
    ("크로와상 생지", "개", "크로와상 붕어빵용 (생지만 매입)"),
    ("팥", "g", "❓크로와상 붕어빵 토핑? 점장 확인"),
    ("초콜릿", "g", "❓초코쿠키용 (다크초콜릿/카카오?)"),
    ("라우치병 주스(딸기)", "개", "매입 완제품"),
    ("라우치병 주스(망고)", "개", ""),
    ("라우치병 주스(오렌지)", "개", ""),
    ("14oz 핫컵", "개", ""),
    ("14oz 콜드컵", "개", ""),
    ("10oz 핫컵", "개", ""),
    ("10oz 콜드컵", "개", ""),
    ("핫컵 뚜껑", "개", ""),
    ("콜드컵 뚜껑", "개", ""),
    ("홀더", "개", ""),
    ("포장병", "개", "딸기라떼·바닐라빈밀크 ICE 포장용"),
    ("드립백 포장지", "개", "드립백 1개당 포장지"),
    ("디저트 봉투", "개", ""),
]

cash_rows = []
for r, (name, unit, memo) in enumerate(purchased, start=2):
    row = r
    pay = "현금" if name in CASH_INGREDIENTS else "카드"
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=unit)
    ws.cell(row=row, column=3, value="매입")
    ws.cell(row=row, column=4, value=pay)
    ws.cell(row=row, column=5, value=UNK)
    ws.cell(row=row, column=6, value=UNK)
    ws.cell(row=row, column=7, value=UNK)
    ws.cell(row=row, column=8, value=f"=IFERROR(G{row}/(E{row}*F{row}),\"\")")
    ws.cell(row=row, column=9, value=EFF_DATE)
    ws.cell(row=row, column=10, value=memo)
    if pay == "현금":
        cash_rows.append(row)

made = [
    ("자몽청", "g", "→ 시트2 서브레시피에 배합 입력"),
    ("레몬청", "g", "→ 시트2"),
    ("패션푸르츠청", "g", "→ 시트2"),
    ("사과청", "g", "→ 시트2"),
    ("도원결의청", "g", "→ 시트2 / ❓복숭아청?"),
    ("딸기청", "g", "→ 시트2"),
    ("오미자청", "g", "→ 시트2 / 오미자티용"),
    ("말차크림", "g", "→ 시트2 / 말차티라미수용"),
    ("흑임자크림", "g", "→ 시트2 (PDF 확정: 1배치 80g)"),
    ("티라미수베이스", "g", "→ 시트2"),
    ("더치원액", "ml", "→ 시트2 / 더치커피·더치라떼·원액 판매 공통"),
    ("드립백", "개", "→ 시트2 / 원두 갈아 1봉씩 포장"),
    ("초코쿠키", "개", "→ 시트2 / 초코쿠키박스 + 구움과자세트"),
    ("바스크 치즈케익", "조각", "→ 시트2 (1판 → N조각)"),
    ("수제 베이글", "개", "→ 시트2"),
    ("수제 스콘", "개", "→ 시트2"),
    ("수제 마들렌", "개", "→ 시트2"),
    ("수제 호두파이", "개", "→ 시트2"),
]
start = ws.max_row + 1
for i, (name, unit, memo) in enumerate(made):
    row = start + i
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=unit)
    ws.cell(row=row, column=3, value="수제")
    ws.cell(row=row, column=4, value="-")
    ws.cell(row=row, column=5, value="-")
    ws.cell(row=row, column=6, value="-")
    ws.cell(row=row, column=7, value="-")
    ws.cell(row=row, column=8, value="(서브레시피 자동계산)")
    ws.cell(row=row, column=9, value="")
    ws.cell(row=row, column=10, value=memo)

# 자유 추가 영역
add_marker_row = ws.max_row + 1
ws.cell(row=add_marker_row, column=1, value="↓ 점장 자유 추가 영역 (빈 행에 입력) ↓")
for c in ws[add_marker_row]:
    c.font = Font(name="맑은 고딕", bold=True, size=10, color="B33A3A")
    c.fill = ADD_FILL
    c.border = BORDER
    c.alignment = LEFT if c.column == 1 else CENTER

for i in range(ADD_ROWS):
    row = add_marker_row + 1 + i
    ws.cell(row=row, column=8, value=f"=IFERROR(G{row}/(E{row}*F{row}),\"\")")
    for col in range(1, 11):
        c = ws.cell(row=row, column=col)
        c.fill = ADD_FILL
        c.border = BORDER

style_header(ws)
style_body(ws, start_row=2, name_col=1, memo_col=10)

for row_idx in cash_rows:
    cell = ws.cell(row=row_idx, column=1)
    cell.fill = CASH_FILL

autosize(ws)
ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════
# 시트 2: 서브레시피 (샘플 가짜 숫자 포함)
# ═══════════════════════════════════════════════
ws = wb.create_sheet("서브레시피")
ws.append(["수제재료명", "1배치 산출량", "구성재료명", "사용량", "비고"])

# (이름, 산출량, 구성, 사용량, 비고, is_sample)
sub = [
    # 자몽청 — 샘플 가짜 숫자
    ("자몽청", 1000, "자몽", 600, "📌 예시 가짜 숫자 — 실제 값으로 교체", True),
    ("자몽청", 1000, "설탕", 400, "📌 예시", True),
    ("레몬청", UNK, "레몬", UNK, "", False),
    ("레몬청", UNK, "설탕", UNK, "", False),
    ("패션푸르츠청", UNK, "패션푸르츠", UNK, "", False),
    ("패션푸르츠청", UNK, "설탕", UNK, "", False),
    ("사과청", UNK, "사과", UNK, "", False),
    ("사과청", UNK, "설탕", UNK, "", False),
    ("도원결의청", UNK, "복숭아", UNK, "❓원물 종류 확인", False),
    ("도원결의청", UNK, "설탕", UNK, "", False),
    ("딸기청", UNK, "딸기", UNK, "", False),
    ("딸기청", UNK, "설탕", UNK, "", False),
    ("오미자청", UNK, "오미자", UNK, "", False),
    ("오미자청", UNK, "설탕", UNK, "", False),
    # 흑임자 — PDF 확정
    ("흑임자크림", 80, "흑임자", 20, "PDF 확정: 흑임자20+생60 (1잔=1배치)", False),
    ("흑임자크림", 80, "생크림", 60, "", False),
    # 말차크림 — 샘플 가짜 숫자
    ("말차크림", 200, "마스카포네", 100, "📌 예시 가짜 숫자 — 실제 값으로 교체", True),
    ("말차크림", 200, "설탕", 50, "📌 예시", True),
    ("말차크림", 200, "말차파우더", 50, "📌 예시", True),
    ("티라미수베이스", UNK, "마스카포네", UNK, "", False),
    ("티라미수베이스", UNK, "설탕", UNK, "", False),
    # 더치 + 드립백
    ("더치원액", UNK, "원두", UNK, "❓1배치 추출량 + 사용 원두 g", False),
    ("드립백", UNK, "원두", UNK, "❓1배치=몇 개, 1봉당 원두 g. A/B/C 등급별 다른지", False),
    ("드립백", UNK, "드립백 포장지", UNK, "", False),
    # 디저트
    ("초코쿠키", UNK, "밀가루", UNK, "❓1배치=몇 개", False),
    ("초코쿠키", UNK, "버터", UNK, "", False),
    ("초코쿠키", UNK, "설탕", UNK, "", False),
    ("초코쿠키", UNK, "초콜릿", UNK, "", False),
    ("초코쿠키", UNK, "계란", UNK, "", False),
    ("바스크 치즈케익", UNK, "크림치즈", UNK, "❓1판 N조각", False),
    ("바스크 치즈케익", UNK, "설탕", UNK, "", False),
    ("바스크 치즈케익", UNK, "계란", UNK, "", False),
    ("바스크 치즈케익", UNK, "밀가루", UNK, "", False),
    ("수제 베이글", UNK, "밀가루", UNK, "❓1배치=N개", False),
    ("수제 베이글", UNK, "버터", UNK, "", False),
    ("수제 스콘", UNK, "밀가루", UNK, "", False),
    ("수제 스콘", UNK, "버터", UNK, "", False),
    ("수제 마들렌", UNK, "밀가루", UNK, "", False),
    ("수제 마들렌", UNK, "계란", UNK, "", False),
    ("수제 마들렌", UNK, "버터", UNK, "", False),
    ("수제 호두파이", UNK, "밀가루", UNK, "", False),
    ("수제 호두파이", UNK, "호두", UNK, "", False),
    ("수제 호두파이", UNK, "버터", UNK, "", False),
]
sample_rows = []
for r in sub:
    name, batch, ing, qty, memo, is_sample = r
    ws.append([name, batch, ing, qty, memo])
    if is_sample:
        sample_rows.append(ws.max_row)

style_header(ws)
style_body(ws, start_row=2, name_col=1, memo_col=5)

# 샘플 행 옅은 파랑 + 회색 글자
for row_idx in sample_rows:
    for c in ws[row_idx]:
        c.fill = SAMPLE_FILL
        c.font = SAMPLE_FONT

autosize(ws)
ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════
# 시트 3: 레시피
# ═══════════════════════════════════════════════
ws = wb.create_sheet("레시피")
ws.append(["메뉴명", "카테고리", "재료명", "사용량", "단위·비고"])

pdf_recipes = {
    "에스프레소": [("원두", 18, "g / 2샷")],
    "아메리카노": [("원두", 18, "g / 2샷, HOT·ICE 통일")],
    "카페라떼": [("원두", 18, "g"), ("우유", 250, "g / HOT 스팀, ICE 차가운 동량")],
    "카푸치노": [("원두", 18, "g / HOT 전용"), ("우유", 210, "g")],
    "리얼바닐라빈 라떼": [("원두", 18, "g"), ("우유", 250, "g"), ("바닐라시럽", 2.5, "펌프")],
    "카라멜마끼아또": [("원두", 18, "g"), ("우유", 250, "g"), ("바닐라시럽", 2.5, "펌프"), ("카라멜드리즐", 1, "회")],
    "카페모카": [("원두", 18, "g"), ("우유", 250, "g"), ("초코시럽", 30, "g")],
    "아인슈페너": [("원두", 18, "g"), ("바닐라시럽", 0.5, "펌프"), ("휘핑크림", UNK, "g"), ("시나몬파우더", 1, "회")],
    "바닐라 슈페너": [("원두", 18, "g"), ("우유", 210, "g"), ("바닐라시럽", 1.5, "펌프"), ("휘핑크림", UNK, "g"), ("시나몬파우더", 1, "회")],
    "흑임자슈페너": [("원두", 18, "g"), ("흑임자크림", 80, "g / 1잔=1배치")],
    "핸드드립": [("핸드드립 원두", UNK, "g")],
    "더치커피(ICE)": [("더치원액", UNK, "ml")],
    "더치라떼(ICE)": [("더치원액", UNK, "ml"), ("우유", UNK, "g")],
    "더치커피 원액260ml": [("더치원액", 260, "ml / 병 판매")],
    "더치커피 원액450ml": [("더치원액", 450, "ml / 병 판매")],
    "드립백 A": [("드립백", 1, "개 / ❓등급별 원두 g 다른지 시트2 확인")],
    "드립백 B": [("드립백", 1, "개 / ❓")],
    "드립백 C": [("드립백", 1, "개 / ❓")],
    "딸기라떼": [("딸기청", 100, "g / ICE 전용, 포장병"), ("우유", 200, "g")],
    "리얼바닐라빈 밀크": [("바닐라시럽", 30, "g"), ("우유", 250, "g / ❓HOT 250 vs ICE 270")],
    "아이스초코": [("초코시럽", 50, "g"), ("우유", 250, "g"), ("코코아파우더", 1, "회"), ("초코드리즐", 1, "회")],
    "핫/아이스초코": [("초코시럽", 50, "g"), ("우유", 250, "g"), ("코코아파우더", 1, "회"), ("초코드리즐", 1, "회")],
    "말차 라떼": [("말차파우더", 10, "g / ❓매장H 기준. 매장I 25g / 포장H 25g 처리"), ("순수", 1, "g"), ("우유", 140, "g / 매장HOT 기준")],
    "순수 말차": [("말차파우더", UNK, "g"), ("순수", UNK, "g")],
    "말차 크림": [("말차크림", UNK, "g"), ("말차파우더", UNK, "g")],
    "말차 딸기 크림(Only Ice)": [("말차크림", UNK, "g"), ("딸기청", UNK, "g"), ("우유", UNK, "g")],
    "도원결의티": [("도원결의청", 30, "g"), ("복숭아시럽", 2, "펌프")],
    "도원결의 에이드": [("도원결의청", 35, "g"), ("복숭아시럽", 2.5, "펌프"), ("탄산수", UNK, "ml")],
    "레몬티": [("레몬청", 50, "g"), ("레몬", 2.5, "g / 슬라이스 2~3"), ("레몬시럽", 1, "펌프 / ❓Hot 1 vs Ice 2")],
    "레몬에이드": [("레몬청", 50, "g"), ("레몬", 2.5, "g"), ("레몬시럽", 2, "펌프"), ("탄산수", UNK, "ml")],
    "자몽티": [("자몽청", 60, "g")],
    "자몽에이드": [("자몽청", 65, "g"), ("탄산수", UNK, "ml")],
    "패션후르츠티": [("패션푸르츠청", 60, "g"), ("슈가시럽", 1.5, "펌프"), ("라임시럽", 0.5, "펌프")],
    "패션후르츠 에이드": [("패션푸르츠청", 75, "g"), ("슈가시럽", 1.5, "펌프"), ("라임시럽", 0.5, "펌프"), ("탄산수", UNK, "ml")],
    "애플시나몬티": [("사과청", 20, "g"), ("사과", 60, "g / 슬라이스")],
    "오미자티": [("오미자청", UNK, "g")],
    "로얄 밀크티": [("로얄밀크티 티백", 2, "개 / 3분 우림"), ("우유", 150, "g"), ("슈가시럽", 4, "펌프 / ❓ICE 미명시")],
    "TWG  레이디그레이티": [("TWG 레이디그레이 티백", 1, "개")],
    "TWG  캐모마일티": [("TWG 캐모마일 티백", 1, "개")],
    "TWG  페퍼민트티": [("TWG 페퍼민트 티백", 1, "개")],
    "팔레데떼 프랑스티 블루오브런던": [("팔레데떼 티백", 1, "개")],
    "아이스티": [("아이스티 베이스", UNK, "g")],
    "망고아이스티": [("아이스티 베이스", UNK, "g"), ("복숭아시럽", UNK, "펌프 / ❓망고시럽 별도?")],
    "복숭아아이스티": [("아이스티 베이스", UNK, "g"), ("복숭아시럽", UNK, "펌프")],
    "봄딸기주스": [("딸기", UNK, "g"), ("설탕", UNK, "g")],
    "이벤트용 봄딸기주스": [("딸기", UNK, "g / ❓봄딸기와 동일?"), ("설탕", UNK, "g")],
    "라우치병 주스(딸기)": [("라우치병 주스(딸기)", 1, "개 / 매입 그대로")],
    "라우치병 주스(망고)": [("라우치병 주스(망고)", 1, "개")],
    "라우치병 주스(오렌지)": [("라우치병 주스(오렌지)", 1, "개")],
    "플레인스콘": [("수제 스콘", 1, "개 / 데우기 10초")],
    "레몬 마들렌": [("수제 마들렌", 1, "개 / 데우기 10초")],
    "미니 마들렌 박스": [("수제 마들렌", UNK, "개 / ❓박스 N개")],
    "플레인 베이글(+크림치즈)": [("수제 베이글", 1, "개 / 데우기 20초"), ("크림치즈", UNK, "g")],
    "잠봉뵈르&고메버터베이글샌드위치": [("수제 베이글", 1, "개"), ("할라피뇨", 6.5, "개"), ("잠봉뵈르햄", 50, "g"), ("버터", 3, "조각")],
    "잠봉뵈르&루꼴라베이글샌드위치": [("수제 베이글", 1, "개"), ("어니언드레싱", 1, "펌프 / 한바퀴"), ("잠봉뵈르햄", UNK, "g / ❓2겹"), ("선드라이토마토", 6.5, "개"), ("발사믹", 1, "펌프"), ("루꼴라", UNK, "g")],
    "말차 티라미수": [("티라미수베이스", 35, "g"), ("말차크림", 110, "g")],
    "말차 티라미수(미니)": [("티라미수베이스", 25, "g"), ("말차크림", 50, "g")],
    "바스크 치즈케익": [("바스크 치즈케익", 1, "조각")],
    "초코 쿠키 박스": [("초코쿠키", UNK, "개 / ❓박스 N개")],
    "호두파이": [("수제 호두파이", 1, "개")],
    "크로와상 붕어빵": [("크로와상 생지", 1, "개"), ("팥", UNK, "g")],
    "구움과자세트": [("초코쿠키", 4, "개"), ("수제 마들렌", 1, "개 / ❓레몬마들렌 1~2"), ("수제 스콘", 1, "개")],
    "베이글샌드위치1종+아메리카노 세트": [("수제 베이글", 1, "개 / ❓잠봉뵈르 가정"), ("잠봉뵈르햄", UNK, "g"), ("원두", 18, "g / 아메리카노")],
    "플레인베이글+아메리카노 세트": [("수제 베이글", 1, "개"), ("크림치즈", UNK, "g"), ("원두", 18, "g / 아메리카노")],
    "교토 말차세트": [("말차파우더", UNK, "g / ❓말차+디저트?"), ("우유", UNK, "g")],
    "순수 말차세트": [("말차파우더", UNK, "g / ❓순수말차+디저트?"), ("순수", UNK, "g")],
}

sorted_products = sorted(
    products,
    key=lambda p: (cat_by_name.get(p["name"], "zzz"), p["name"]),
)

unmatched_count = 0
for p in sorted_products:
    name = p["name"]
    cat = cat_by_name.get(name, "")
    if name in pdf_recipes:
        for ing, qty, memo in pdf_recipes[name]:
            ws.append([name, cat, ing, qty, memo])
    else:
        ws.append([name, cat, UNK, UNK, "❓PDF 미수록 — 점장이 레시피 입력"])
        unmatched_count += 1

style_header(ws)
style_body(ws, start_row=2, name_col=1, memo_col=5)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[2].value == UNK and "PDF 미수록" in (row[4].value or ""):
        for c in row:
            if c.value != UNK:
                c.fill = NEW_MENU_FILL

autosize(ws)
ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════
# 시트 4: 포장세트
# ═══════════════════════════════════════════════
ws = wb.create_sheet("포장세트")
ws.append(["세트명", "카테고리", "온도", "재료명", "사용량", "비고"])

packs = [
    ("커피 HOT 포장", "coffee", "hot", "14oz 핫컵", 1, ""),
    ("커피 HOT 포장", "coffee", "hot", "핫컵 뚜껑", 1, ""),
    ("커피 HOT 포장", "coffee", "hot", "홀더", 1, ""),
    ("커피 ICE 포장", "coffee", "ice", "14oz 콜드컵", 1, ""),
    ("커피 ICE 포장", "coffee", "ice", "콜드컵 뚜껑", 1, ""),
    ("커피 ICE 포장", "coffee", "ice", "홀더", 1, ""),
    ("말차 HOT 포장", "matcha", "hot", "14oz 핫컵", 1, ""),
    ("말차 HOT 포장", "matcha", "hot", "핫컵 뚜껑", 1, ""),
    ("말차 HOT 포장", "matcha", "hot", "홀더", 1, ""),
    ("말차 ICE 포장", "matcha", "ice", "14oz 콜드컵", 1, ""),
    ("말차 ICE 포장", "matcha", "ice", "콜드컵 뚜껑", 1, ""),
    ("말차 ICE 포장", "matcha", "ice", "홀더", 1, ""),
    ("티 HOT 포장", "tea", "hot", "14oz 핫컵", 1, ""),
    ("티 HOT 포장", "tea", "hot", "핫컵 뚜껑", 1, ""),
    ("티 HOT 포장", "tea", "hot", "홀더", 1, ""),
    ("티 ICE 포장", "tea", "ice", "14oz 콜드컵", 1, ""),
    ("티 ICE 포장", "tea", "ice", "콜드컵 뚜껑", 1, ""),
    ("티 ICE 포장", "tea", "ice", "홀더", 1, ""),
    ("에이드 ICE 포장", "ade", "ice", "14oz 콜드컵", 1, ""),
    ("에이드 ICE 포장", "ade", "ice", "콜드컵 뚜껑", 1, ""),
    ("에이드 ICE 포장", "ade", "ice", "홀더", 1, ""),
    ("음료 ICE 포장", "beverage", "ice", "포장병", 1, "❓딸기라떼/바닐라빈밀크 기준"),
    ("드립커피 HOT 포장", "drip_coffee", "hot", UNK, UNK, "❓핸드드립 HOT 컵 종류"),
    ("드립커피 ICE 포장", "drip_coffee", "ice", UNK, UNK, "❓"),
    ("더치커피 ICE 포장", "dutch_coffee", "ice", UNK, UNK, "❓더치커피(ICE) 컵 / 원액병은 별도"),
    ("시즌 포장", "season", "", UNK, UNK, "❓봄딸기주스 컵"),
    ("디저트 포장", "dessert", "", "디저트 봉투", 1, "❓봉투 종류"),
]
for p in packs:
    ws.append(list(p))
style_header(ws)
style_body(ws, start_row=2, name_col=1, memo_col=6)
autosize(ws)
ws.freeze_panes = "A2"


wb.save(OUT)
print(f"saved: {OUT}")
print(f"PDF/추정 매칭 메뉴: {len(pdf_recipes)}개 / 활성 메뉴 총 {len(products)}개 / 미매칭 {unmatched_count}개")
print(f"매입재료: {len(purchased)}개 / 수제재료: {len(made)}개 / 현금그룹: {len(cash_rows)}개")
