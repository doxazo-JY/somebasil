"""테스트용 — v5에서 원두만 채워서 업로드 테스트.

원두 1kg봉 × 10봉 = 350,000원 → 단가 35원/g 자동
다른 재료는 ❓ 그대로 둠.
서브레시피 시트의 샘플 가짜 행(자몽청/말차크림)은 깨끗이 비움.

기대 결과:
- 아메리카노, 에스프레소: status='ok' (원두만 쓰니까)
- 나머지: 'missing_price' or 'no_recipe'
"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

SRC = r"C:\Users\PC\Desktop\somebasil\notes\레시피_템플릿_v5.xlsx"
DST = r"C:\Users\PC\Desktop\somebasil\notes\레시피_템플릿_v5_test_원두.xlsx"
UNK = "❓"
UNK_FILL = PatternFill("solid", fgColor="FFE066")
PLAIN_FONT = Font(name="맑은 고딕", size=10)

wb = load_workbook(SRC)

# ─── 시트1 재료: 원두 행 채우기 ───
ws = wb["재료"]
filled = False
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].value == "원두":
        row[4].value = 1000
        row[5].value = 10
        row[6].value = 350000
        filled = True
        print(f"row {row[0].row}: 원두 → 1000g × 10봉 × 350,000원 = 35원/g")
        break
if not filled:
    print("⚠ '원두' 행 못 찾음")

# ─── 시트2 서브레시피: 샘플 가짜 행 깨끗이 비우기 ───
ws2 = wb["서브레시피"]
cleared = 0
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    memo = str(row[4].value or "")
    if "📌" in memo or "예시" in memo:
        # 산출량/사용량을 ❓ 으로, 메모 비움
        row[1].value = UNK
        row[1].fill = UNK_FILL
        row[1].font = PLAIN_FONT
        row[3].value = UNK
        row[3].fill = UNK_FILL
        row[3].font = PLAIN_FONT
        row[4].value = ""
        row[4].font = PLAIN_FONT
        # 다른 셀의 샘플 배경도 정리
        for c in row:
            if c.column in (2, 4):
                c.font = PLAIN_FONT
        cleared += 1
print(f"샘플 행 정리: {cleared}개")

wb.save(DST)
print(f"saved: {DST}")
