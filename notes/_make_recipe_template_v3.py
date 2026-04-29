"""레시피 템플릿 v3 — v2 + 결제방식 컬럼.

변경점 (v2 → v3):
- 시트1 재료에 '결제방식'(현금/카드) 컬럼 추가
  - 현금 4종(원두·우유·말차파우더·잠봉뵈르햄): 거래 단위 추적 가능
  - 카드: 묶음 결제라 월 단위 합계로만 비교
- README에 1·2·3·4 계산 흐름 그림 추가
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\PC\Desktop\somebasil\notes\레시피_템플릿_v3.xlsx"
MENU_DATA = r"C:\Users\PC\Desktop\somebasil\notes\_menu_data.json"
UNK = "❓"
EFF_DATE = "2025-12-15"

# 현금 결제 4종 (홍인호·한성욱·김인성·소금집)
CASH_INGREDIENTS = {"원두", "우유", "말차파우더", "잠봉뵈르햄"}

with open(MENU_DATA, encoding="utf-8") as f:
    md = json.load(f)
products = md["products"]
cat_by_name = md["category_by_name"]

# ─── 스타일 ───
HEAD_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
HEAD_FILL = PatternFill("solid", fgColor="1A5C3A")
BODY_FONT = Font(name="맑은 고딕", size=10)
UNK_FILL = PatternFill("solid", fgColor="FFF3C4")
NEW_MENU_FILL = PatternFill("solid", fgColor="E8F1ED")
CASH_FILL = PatternFill("solid", fgColor="E8F1ED")  # 현금 4종 행 강조
THIN = Side(border_style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def style_body(ws, start_row=2, name_col=1):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = LEFT if cell.column == name_col else CENTER
            if cell.value == UNK:
                cell.fill = UNK_FILL

def autosize(ws, padding=4, max_w=50):
    for col in ws.columns:
        m = 0
        letter = get_column_letter(col[0].column)
        for c in col:
            v = "" if c.value is None else str(c.value)
            l = sum(2 if ord(ch) > 127 else 1 for ch in v)
            if l > m: m = l
        ws.column_dimensions[letter].width = min(m + padding, max_w)


wb = Workbook()

# ═══════════════════════════════════════════════
# 시트 0: README
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = "README"
ws["A1"] = "썸바실 레시피·원가 입력 양식 (v3 — 결제방식 추가)"
ws["A1"].font = Font(name="맑은 고딕", bold=True, size=15, color="1A5C3A")
ws.row_dimensions[1].height = 28

guide = [
    "",
    "이 파일은 PDF 레시피정리(2026.02) + DB 활성메뉴 66개를 합쳐 자동 생성된 초안입니다.",
    "",
    "■ 시트 구성",
    "  1. 재료 — 매입재료/수제재료 마스터. 단가는 [총가격÷(용량×세트수)] 자동계산",
    "  2. 서브레시피 — 수제재료(청·크림·베이커리) 1배치 배합",
    "  3. 레시피 — DB 활성 메뉴 66개 전체 + PDF에서 매칭된 사용량",
    "  4. 포장세트 — 카테고리(영문) × 온도(hot/ice)별 포장재 묶음",
    "",
    "■ v3 추가 — '결제방식' 컬럼",
    "  • 현금: 통장에 거래 단위로 찍히는 재료 (원두·우유·말차파우더·잠봉뵈르햄 4종)",
    "  • 카드: 쿠팡·이마트·씨유 등 묶음 결제라 시점별 추적 불가",
    "  • 활용: 4번 화면에서 '봉지 단위 회수율'(현금) vs '월 단위 합계 비교'(카드)로 분리 표시",
    "",
    "■ 색깔 의미",
    "  • 노란색 (❓) : PDF·DB 어디에도 없어 점장 입력 필요",
    "  • 연두색 행 : DB에는 있는데 PDF에 없는 메뉴 — 점장이 레시피 채워야 함",
    "  • 시트1 연두색 행 : 현금 결제 4종 — 거래 추적 가능 그룹",
    "",
    "─────────────────────────────────────────────",
    "■ 이 엑셀이 어떻게 활용되는지 (계산 흐름)",
    "─────────────────────────────────────────────",
    "",
    "1. 메뉴별 원가 계산",
    "   메뉴원가 = Σ(시트3 레시피 사용량 × 시트1 단가) + 시트4 포장세트 비용",
    "   예) 아메리카노 = 원두 18g × 35원 + 컵세트 280원 = 910원",
    "",
    "2. 메뉴별 마진 / 원가율",
    "   잔당 마진 = DB 판매가(products.price) − 메뉴원가",
    "   예) 아메 4,500원 − 910원 = 3,590원 (원가율 20%)",
    "",
    "3. 재료 1세트 회수 분석 (실제 mix 기반)",
    "   봉지 며칠치 = 1세트 용량 ÷ 일평균 사용량",
    "   일평균 사용량 = Σ(daily_sales 일판매잔수 × 잔당 사용량)",
    "   N일치 마진 합 = Σ(같은 N일간 그 재료 쓴 메뉴들의 잔당 마진 × 판매잔수)",
    "   회수율 = N일치 마진 ÷ 봉지 가격",
    "   예) 원두 1봉(35,000원) = 평균 4.2일치 → 그 기간 커피류 마진 70만 (회수율 20배)",
    "",
    "4. 현 상황 진단 (월 단위)",
    "   [현금 4종] 거래 단위 추적",
    "     monthly_expenses에서 counterpart로 필터 → 봉지별 매칭",
    "     '4월 원두 70만 구매 → 추정 사용 65만 (재고 5만)'",
    "   [카드 묶음] 월 단위 합계만",
    "     '4월 카드 재료비 합 100만 vs 매출 추정 사용분 90만'",
    "",
    "─────────────────────────────────────────────",
    "■ 시트1 재료 입력 방법 (예시)",
    "─────────────────────────────────────────────",
    "  원두 1kg짜리 봉지 2개를 70,000원에 샀다면:",
    "    1세트 용량 = 1000, 세트 수 = 2, 총 가격 = 70000",
    "    → 단가 = 35원/g 자동계산",
    "  바닐라시럽 한 병이 200펌프 나오고 30,000원이면:",
    "    1세트 용량 = 200(펌프), 세트 수 = 1, 총 가격 = 30000",
    "    → 단가 = 150원/펌프 자동계산",
    "  ※ '용량' 단위는 레시피에서 쓰는 단위와 같아야 함. 환산은 점장이 미리.",
    "",
    "■ 적용시작일",
    "  현재 모든 행 = 2025-12-15 통일 (시스템 도입 기준일)",
    "  추후 가격 변동 시 새 행 추가 + 그 변경일 입력",
    "",
    "■ 점장 컨펌 핵심 항목",
    "  • 시트1 모든 매입재료의 (1세트 용량 / 세트 수 / 총가격) 입력",
    "  • 시트1 결제방식(현금/카드) 확인 — 사전입력 외 추가 수정",
    "  • 시트2 수제재료별 1배치 산출량 + 구성재료 (자몽청, 말차크림, 베이커리류 등)",
    "  • 시트3 PDF 미수록 메뉴(연두색)의 레시피 — 35개",
    "  • 시트4 컵 종류별 단가 (시트1에 등록된 컵명과 일치)",
    "  • 말차라떼 매장/포장 사용량 차이 처리",
    "  • '순수'(말차에 들어가는 것)의 정체",
    "  • 호두파이/얼그레이 3종 — DB 매칭 메뉴 확인 필요",
]
for i, line in enumerate(guide, start=2):
    c = ws.cell(row=i, column=1, value=line)
    c.font = Font(name="맑은 고딕", size=11)
ws.column_dimensions["A"].width = 100


# ═══════════════════════════════════════════════
# 시트 1: 재료 (구매현황 → 단가 자동계산 + 결제방식)
# ═══════════════════════════════════════════════
ws = wb.create_sheet("재료")
ws.append([
    "재료명", "사용단위", "종류", "결제방식",
    "1세트 용량", "세트 수", "총 가격(원)", "단가(자동)",
    "적용시작일", "비고",
])

# (재료명, 사용단위, 비고)
purchased = [
    # 커피·라떼 (현금 그룹)
    ("원두", "g", "예: 1kg봉지×2개=70000원 → 1세트용량 1000, 세트수 2"),
    ("우유", "g", "PDF 기준 g (스팀 우유 무게)"),
    # 시럽·소스
    ("바닐라시럽", "펌프", "예: 1병=200펌프 환산"),
    ("초코시럽", "g", ""),
    ("카라멜드리즐", "회", "토핑 1회분"),
    ("시나몬파우더", "회", "토핑"),
    ("코코아파우더", "회", "토핑"),
    ("초코드리즐", "회", "토핑"),
    ("휘핑크림", "g", ""),
    # 흑임자크림 구성
    ("흑임자", "g", ""),
    ("생크림", "g", ""),
    # 말차
    ("말차파우더", "g", ""),
    ("순수", "g", "❓PDF '순수'의 정체 확인 필요"),
    # 시럽 (펌프)
    ("복숭아시럽", "펌프", ""),
    ("레몬시럽", "펌프", ""),
    ("슈가시럽", "펌프", ""),
    ("라임시럽", "펌프", ""),
    # 차
    ("얼그레이 티백", "개", "3종 — 종류별 단가 다르면 분리"),
    ("로얄밀크티 티백", "개", ""),
    ("탄산수", "ml", ""),
    # 청 원물
    ("자몽", "g", ""),
    ("레몬", "g", "원물 + 가니쉬용"),
    ("사과", "g", "청 + 슬라이스"),
    ("복숭아", "g", "❓도원결의청용 추정"),
    ("패션푸르츠", "g", ""),
    ("딸기", "g", ""),
    ("설탕", "g", "청 + 베이커리"),
    # 잠봉뵈르 (햄은 현금 그룹)
    ("잠봉뵈르햄", "g", ""),
    ("버터", "조각", "PDF: 3조각"),
    ("할라피뇨", "개", ""),
    ("어니언드레싱", "펌프", "PDF '한바퀴' = 1펌프 추정"),
    ("선드라이토마토", "개", ""),
    ("발사믹", "펌프", ""),
    ("루꼴라", "g", ""),
    # 베이커리
    ("마스카포네", "g", "❓티라미수베이스 추정"),
    ("밀가루", "g", "❓베이커리 자체제작"),
    ("계란", "개", "❓"),
    ("호두", "g", "❓"),
    # 포장재
    ("14oz 핫컵", "개", ""),
    ("14oz 콜드컵", "개", ""),
    ("10oz 핫컵", "개", ""),
    ("10oz 콜드컵", "개", ""),
    ("핫컵 뚜껑", "개", ""),
    ("콜드컵 뚜껑", "개", ""),
    ("홀더", "개", ""),
    ("포장병", "개", "딸기라떼·바닐라빈밀크 ICE 포장용"),
    ("디저트 봉투", "개", ""),
]

cash_rows = []
for r, (name, unit, memo) in enumerate(purchased, start=2):
    row = r
    pay = "현금" if name in CASH_INGREDIENTS else "카드"
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=unit)
    ws.cell(row=row, column=3, value="매입")
    ws.cell(row=row, column=4, value=pay)
    ws.cell(row=row, column=5, value=UNK)  # 1세트 용량
    ws.cell(row=row, column=6, value=UNK)  # 세트 수
    ws.cell(row=row, column=7, value=UNK)  # 총가격
    # 단가 수식: =G/(E*F)
    ws.cell(row=row, column=8, value=f"=IFERROR(G{row}/(E{row}*F{row}),\"\")")
    ws.cell(row=row, column=9, value=EFF_DATE)
    ws.cell(row=row, column=10, value=memo)
    if pay == "현금":
        cash_rows.append(row)

# 수제 재료
made = [
    ("자몽청", "g", ""),
    ("레몬청", "g", ""),
    ("패션푸르츠청", "g", ""),
    ("사과청", "g", ""),
    ("도원결의청", "g", "❓복숭아청? 점장 확인"),
    ("딸기청", "g", ""),
    ("말차크림", "g", "말차티라미수용"),
    ("흑임자크림", "g", "PDF: 흑임자20+생60 = 80g 1배치 (확정)"),
    ("티라미수베이스", "g", ""),
    ("수제 베이글", "개", ""),
    ("수제 스콘", "개", ""),
    ("수제 마들렌", "개", ""),
    ("수제 호두파이", "개", "❓DB에 호두파이 단품 없음 — 메뉴 매칭 확인"),
]
start = ws.max_row + 1
for i, (name, unit, memo) in enumerate(made):
    row = start + i
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=unit)
    ws.cell(row=row, column=3, value="수제")
    ws.cell(row=row, column=4, value="-")
    ws.cell(row=row, column=5, value="-")
    ws.cell(row=row, column=6, value="-")
    ws.cell(row=row, column=7, value="-")
    ws.cell(row=row, column=8, value="(서브레시피 자동계산)")
    ws.cell(row=row, column=9, value="")
    ws.cell(row=row, column=10, value=memo)

style_header(ws)
style_body(ws)

# 현금 4종 행 연두색 강조
for row_idx in cash_rows:
    for c in ws[row_idx]:
        if c.value not in (UNK, None) and not (isinstance(c.value, str) and c.value.startswith("=")):
            c.fill = CASH_FILL

autosize(ws)
ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════
# 시트 2: 서브레시피
# ═══════════════════════════════════════════════
ws = wb.create_sheet("서브레시피")
ws.append(["수제재료명", "1배치 산출량", "구성재료명", "사용량", "비고"])

sub = [
    ("흑임자크림", 80, "흑임자", 20, "PDF: 흑임자20+생60 (1잔 분량 = 1배치)"),
    ("흑임자크림", 80, "생크림", 60, ""),
    ("자몽청", UNK, "자몽", UNK, "❓1배치 산출량·구성비"),
    ("자몽청", UNK, "설탕", UNK, ""),
    ("레몬청", UNK, "레몬", UNK, "❓"),
    ("레몬청", UNK, "설탕", UNK, ""),
    ("패션푸르츠청", UNK, "패션푸르츠", UNK, "❓"),
    ("패션푸르츠청", UNK, "설탕", UNK, ""),
    ("사과청", UNK, "사과", UNK, "❓"),
    ("사과청", UNK, "설탕", UNK, ""),
    ("도원결의청", UNK, "복숭아", UNK, "❓원물 종류 확인"),
    ("도원결의청", UNK, "설탕", UNK, ""),
    ("딸기청", UNK, "딸기", UNK, "❓"),
    ("딸기청", UNK, "설탕", UNK, ""),
    ("말차크림", UNK, "마스카포네", UNK, "❓구성·산출량"),
    ("말차크림", UNK, "설탕", UNK, ""),
    ("말차크림", UNK, "말차파우더", UNK, ""),
    ("티라미수베이스", UNK, "마스카포네", UNK, "❓"),
    ("티라미수베이스", UNK, "설탕", UNK, ""),
    ("수제 베이글", UNK, "밀가루", UNK, "❓1배치 = N개?"),
    ("수제 베이글", UNK, "버터", UNK, ""),
    ("수제 스콘", UNK, "밀가루", UNK, "❓"),
    ("수제 스콘", UNK, "버터", UNK, ""),
    ("수제 마들렌", UNK, "밀가루", UNK, "❓"),
    ("수제 마들렌", UNK, "계란", UNK, ""),
    ("수제 마들렌", UNK, "버터", UNK, ""),
    ("수제 호두파이", UNK, "밀가루", UNK, "❓"),
    ("수제 호두파이", UNK, "호두", UNK, ""),
    ("수제 호두파이", UNK, "버터", UNK, ""),
]
for r in sub:
    ws.append(list(r))
style_header(ws)
style_body(ws)
autosize(ws)
ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════
# 시트 3: 레시피 (DB 활성 메뉴 66개 + PDF 매칭)
# ═══════════════════════════════════════════════
ws = wb.create_sheet("레시피")
ws.append(["메뉴명", "카테고리", "재료명", "사용량", "비고"])

pdf_recipes = {
    "에스프레소": [("원두", 18, "2샷")],
    "아메리카노": [("원두", 18, "2샷 / HOT·ICE 통일 (물·얼음 무시)")],
    "카페라떼": [
        ("원두", 18, ""),
        ("우유", 250, "HOT 스팀 / ICE 동량 차가운 우유"),
    ],
    "카푸치노": [
        ("원두", 18, "HOT 전용"),
        ("우유", 210, ""),
    ],
    "리얼바닐라빈 라떼": [
        ("원두", 18, ""),
        ("우유", 250, ""),
        ("바닐라시럽", 2.5, ""),
    ],
    "카라멜마끼아또": [
        ("원두", 18, ""),
        ("우유", 250, ""),
        ("바닐라시럽", 2.5, ""),
        ("카라멜드리즐", 1, ""),
    ],
    "카페모카": [
        ("원두", 18, ""),
        ("우유", 250, ""),
        ("초코시럽", 30, ""),
    ],
    "아인슈페너": [
        ("원두", 18, ""),
        ("바닐라시럽", 0.5, ""),
        ("휘핑크림", UNK, "❓잔당 g"),
        ("시나몬파우더", 1, ""),
    ],
    "바닐라 슈페너": [
        ("원두", 18, ""),
        ("우유", 210, ""),
        ("바닐라시럽", 1.5, ""),
        ("휘핑크림", UNK, "❓"),
        ("시나몬파우더", 1, ""),
    ],
    "딸기라떼": [
        ("딸기청", 100, "ICE 전용, 포장병"),
        ("우유", 200, ""),
    ],
    "리얼바닐라빈 밀크": [
        ("바닐라시럽", 30, ""),
        ("우유", 250, "HOT 250 / ICE 270 — 통일 시 점장 확인"),
    ],
    "아이스초코": [
        ("초코시럽", 50, ""),
        ("우유", 250, ""),
        ("코코아파우더", 1, ""),
        ("초코드리즐", 1, ""),
    ],
    "핫/아이스초코": [
        ("초코시럽", 50, ""),
        ("우유", 250, ""),
        ("코코아파우더", 1, ""),
        ("초코드리즐", 1, ""),
    ],
    "말차 라떼": [
        ("말차파우더", 10, "❓PDF 매장 H 기준. 매장 I 25g / 포장 H 25g 처리 정책 확인"),
        ("순수", 1, "❓"),
        ("우유", 140, "❓매장 HOT 기준"),
    ],
    "도원결의티": [
        ("도원결의청", 30, ""),
        ("복숭아시럽", 2, ""),
    ],
    "도원결의 에이드": [
        ("도원결의청", 35, ""),
        ("복숭아시럽", 2.5, ""),
        ("탄산수", UNK, "❓잔당 ml"),
    ],
    "레몬티": [
        ("레몬청", 50, ""),
        ("레몬", 2.5, "PDF: 2~3"),
        ("레몬시럽", 1, "Hot 1펌프 / Ice 2펌프 — 통합 시 점장 확인"),
    ],
    "레몬에이드": [
        ("레몬청", 50, ""),
        ("레몬", 2.5, ""),
        ("레몬시럽", 2, ""),
        ("탄산수", UNK, "❓"),
    ],
    "자몽티": [("자몽청", 60, "")],
    "자몽에이드": [
        ("자몽청", 65, ""),
        ("탄산수", UNK, "❓"),
    ],
    "패션후르츠티": [
        ("패션푸르츠청", 60, ""),
        ("슈가시럽", 1.5, ""),
        ("라임시럽", 0.5, ""),
    ],
    "패션후르츠 에이드": [
        ("패션푸르츠청", 75, ""),
        ("슈가시럽", 1.5, ""),
        ("라임시럽", 0.5, ""),
        ("탄산수", UNK, "❓"),
    ],
    "애플시나몬티": [
        ("사과청", 20, ""),
        ("사과", 60, "슬라이스"),
    ],
    "로얄 밀크티": [
        ("로얄밀크티 티백", 2, "3분 우림"),
        ("우유", 150, ""),
        ("슈가시럽", 4, "HOT만? ICE는 PDF 미명시"),
    ],
    "플레인스콘": [("수제 스콘", 1, "데우기 10초")],
    "레몬 마들렌": [("수제 마들렌", 1, "데우기 10초")],
    "플레인 베이글(+크림치즈)": [("수제 베이글", 1, "데우기 20초 + 크림치즈")],
    "잠봉뵈르&고메버터베이글샌드위치": [
        ("수제 베이글", 1, ""),
        ("할라피뇨", 6.5, ""),
        ("잠봉뵈르햄", 50, ""),
        ("버터", 3, "조각"),
    ],
    "잠봉뵈르&루꼴라베이글샌드위치": [
        ("수제 베이글", 1, ""),
        ("어니언드레싱", 1, "한바퀴 = 1펌프"),
        ("잠봉뵈르햄", UNK, "❓2겹 → g"),
        ("선드라이토마토", 6.5, ""),
        ("발사믹", 1, ""),
        ("루꼴라", UNK, "❓잔당 g"),
    ],
    "말차 티라미수": [
        ("티라미수베이스", 35, ""),
        ("말차크림", 110, ""),
    ],
    "말차 티라미수(미니)": [
        ("티라미수베이스", 25, ""),
        ("말차크림", 50, ""),
    ],
}

sorted_products = sorted(
    products,
    key=lambda p: (cat_by_name.get(p["name"], "zzz"), p["name"]),
)

unmatched_count = 0
for p in sorted_products:
    name = p["name"]
    cat = cat_by_name.get(name, "")
    if name in pdf_recipes:
        for ing, qty, memo in pdf_recipes[name]:
            ws.append([name, cat, ing, qty, memo])
    else:
        ws.append([name, cat, UNK, UNK, "❓PDF 미수록 — 점장이 레시피 입력"])
        unmatched_count += 1

style_header(ws)
style_body(ws)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[2].value == UNK and "PDF 미수록" in (row[4].value or ""):
        for c in row:
            if c.value != UNK:
                c.fill = NEW_MENU_FILL

autosize(ws)
ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════
# 시트 4: 포장세트
# ═══════════════════════════════════════════════
ws = wb.create_sheet("포장세트")
ws.append(["세트명", "카테고리", "온도", "재료명", "사용량", "비고"])

packs = [
    ("커피 HOT 포장", "coffee", "hot", "14oz 핫컵", 1, ""),
    ("커피 HOT 포장", "coffee", "hot", "핫컵 뚜껑", 1, ""),
    ("커피 HOT 포장", "coffee", "hot", "홀더", 1, ""),
    ("커피 ICE 포장", "coffee", "ice", "14oz 콜드컵", 1, ""),
    ("커피 ICE 포장", "coffee", "ice", "콜드컵 뚜껑", 1, ""),
    ("커피 ICE 포장", "coffee", "ice", "홀더", 1, ""),
    ("말차 HOT 포장", "matcha", "hot", "14oz 핫컵", 1, ""),
    ("말차 HOT 포장", "matcha", "hot", "핫컵 뚜껑", 1, ""),
    ("말차 HOT 포장", "matcha", "hot", "홀더", 1, ""),
    ("말차 ICE 포장", "matcha", "ice", "14oz 콜드컵", 1, ""),
    ("말차 ICE 포장", "matcha", "ice", "콜드컵 뚜껑", 1, ""),
    ("말차 ICE 포장", "matcha", "ice", "홀더", 1, ""),
    ("티 HOT 포장", "tea", "hot", "14oz 핫컵", 1, ""),
    ("티 HOT 포장", "tea", "hot", "핫컵 뚜껑", 1, ""),
    ("티 HOT 포장", "tea", "hot", "홀더", 1, ""),
    ("티 ICE 포장", "tea", "ice", "14oz 콜드컵", 1, ""),
    ("티 ICE 포장", "tea", "ice", "콜드컵 뚜껑", 1, ""),
    ("티 ICE 포장", "tea", "ice", "홀더", 1, ""),
    ("에이드 ICE 포장", "ade", "ice", "14oz 콜드컵", 1, ""),
    ("에이드 ICE 포장", "ade", "ice", "콜드컵 뚜껑", 1, ""),
    ("에이드 ICE 포장", "ade", "ice", "홀더", 1, ""),
    ("음료 ICE 포장", "beverage", "ice", "포장병", 1, "❓딸기라떼/바닐라빈밀크 기준"),
    ("드립커피 HOT 포장", "drip_coffee", "hot", UNK, UNK, "❓PDF 미수록"),
    ("드립커피 ICE 포장", "drip_coffee", "ice", UNK, UNK, "❓PDF 미수록"),
    ("더치커피 ICE 포장", "dutch_coffee", "ice", UNK, UNK, "❓"),
    ("시즌 포장", "season", "", UNK, UNK, "❓"),
    ("디저트 포장", "dessert", "", "디저트 봉투", 1, "❓봉투 종류"),
]
for p in packs:
    ws.append(list(p))
style_header(ws)
style_body(ws)
autosize(ws)
ws.freeze_panes = "A2"


wb.save(OUT)
print(f"saved: {OUT}")
print(f"PDF 매칭 메뉴: {len(pdf_recipes)}개 / 활성 메뉴 총 {len(products)}개 / 미매칭 {unmatched_count}개")
print(f"현금 그룹 재료: {len(cash_rows)}개")
